# genetic programming for Gymnasium Pendulum task
# https://www.gymlibrary.dev/environments/classic_control/pendulum/ 

import numpy
import random
import gymnasium as gym
import operator
import matplotlib.pyplot as plt
import math
from openpyxl import load_workbook
import sys
import inspect
from inspect import isclass
from operator import attrgetter

from deap import algorithms
from deap import base
from deap import creator
from deap import tools
from deap import gp

import pygraphviz as pgv

# parallel
import multiprocessing

# Import modules from different directory
import os
import sys
PATH=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(PATH)

# from modules.prim_functions import conditional, truncate
from modules.output_functions import get_one_column
# from modules.eval_individual import fullObsEvalIndividual

NUM_EP=20
LIN_MIN=1
LIN_MAX=13
MAX_STEP=300
GRAV=-9.81
POP=50
PROCESSES=2
GENS=25

def write_to_excel(fit_mins, sheet_name, path):
    workbook = load_workbook(filename=path)

    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        workbook.active=workbook[sheet_name]
        workbook.active.append(['ind', 'fitness'])


    workbook.active=workbook[sheet_name]

    workbook.active.append(fit_mins)

    workbook.save(filename=path)

def truncate(number, decimals=0):
    if math.isinf(number) or math.isnan(number):
        return 0
    if not isinstance(decimals, int):
        raise TypeError("decimal places must be an integer.")
    elif decimals < 0:
        raise ValueError("decimal places has to be 0 or more.")
    elif decimals == 0:
        return math.trunc(number)

def conditional(input1, input2):
    if input1 < input2:
        return -input1
    else: return input1

def fullObsEvalIndividual(individual, pset, grav, test=False):
    # Set up the enviornment and gravity
    num_episode = NUM_EP
    # gravs = numpy.linspace(LIN_MIN, LIN_MAX, num_episode)

    if test:
        current_env = gym.make('Pendulum-v1', g=grav, render_mode="human") # For rendering
        env = current_env
        num_episode = 1
    
    # Transform the tree expression to functional Python code
    get_action = gp.compile(individual, pset)
    fitness = 0
    failed = False
    for x in range(0, num_episode):
        if x > 0:
            gravity=9.81
        else:
            gravity=-9.81
        # Set up the variables for the env
        if not test:
            current_env = gym.make('Pendulum-v1', g=gravity) # For training

        env = current_env
        done = False
        truncated = False
        observation = env.reset()
        observation = observation[0]
        episode_reward = 0
        num_steps = 0
        max_steps=MAX_STEP
        timeout=False

        while not (done or timeout):
            if failed:
                action = 0
            else:
                # use the tree to compute action, plugs values of observation into get_action
                action = get_action(observation[0], observation[1], observation[2])
                action = (action,)

            try: observation, reward, done, truncated, info = env.step(action) # env.step will return the new observation, reward, done, truncated, info
            except:
                failed = True
                observation, reward, done, truncated, info = env.step(0)

            print(observation[0], observation[1])
            # print(math.tan((observation[0], observation[1])))

            if gravity < 0:
                reward=reward*-1
            episode_reward += reward

            num_steps += 1
            if num_steps >= max_steps:
                timeout=True

        fitness += episode_reward
    fitness = fitness/num_episode      
    return (0,) if failed else (fitness,)

def save_graph(gen, fit_mins, best_fit):

    # Simply change the lines in quottation above to change the values you want to graph
    # Allows you to create multiple plots in one figure

    (fig, ax1) = plt.subplots()
    # Plots using gen as x value and fit_mins as y, both are list
    line1 = ax1.plot(
        gen, fit_mins, 'b-', label="Maximum Fitness"
    )
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Fitness", color="b")
    for (
        tl
    ) in ax1.get_yticklabels():  # Changes colour of ticks and numbers on axis
        tl.set_color("b")

    lns = line1  # lns is a list containing both lines [line1, line2]
    # labs contains the labels of each line (Minimum Fitness and Average Size)
    labs = [l.get_label() for l in lns]
    ax1.legend(lns, labs, loc="lower right")  # Adds then a legend

    plt.axis([min(gen), max(gen), min(fit_mins), 0])
    plt.savefig(str(best_fit) + "_fit_curve.png")

# Set up primitives and terminals
pset = gp.PrimitiveSet("MAIN", 3)
pset.addPrimitive(operator.add, 2)
pset.addPrimitive(conditional, 2)

pset.renameArguments(ARG0='y')
pset.renameArguments(ARG1='x')
pset.renameArguments(ARG2='vel')

# Prepare individual and mountain car
creator.create("FitnessMax", base.Fitness, weights=(1.0,))
creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)

toolbox = base.Toolbox()
toolbox.register("expr", gp.genFull, pset=pset, min_=2, max_=3)
toolbox.register("individual", tools.initIterate, creator.Individual,
                 toolbox.expr)
toolbox.register("population", tools.initRepeat, list, toolbox.individual)

# Register functions in the toolbox needed for evolution
toolbox.register("evaluate", fullObsEvalIndividual, pset=pset, grav=GRAV)
toolbox.register("select", tools.selTournament, tournsize=3)
toolbox.register("mate", gp.cxOnePoint)
toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=17))
toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=17))

def main():
    # Initialize the population
    pop = toolbox.population(n=POP)
    hof = tools.HallOfFame(1)

    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", numpy.mean)
    mstats.register("std", numpy.std)
    mstats.register("min", numpy.min)
    mstats.register("max", numpy.max)

    pool = multiprocessing.Pool(processes=PROCESSES) # parllel (Process Pool of 16 workers)
    toolbox.register("map", pool.map) # parallel

    # pop, log = algorithms.eaSimple(pop, toolbox, 0.2, 0.5, GENS, stats=mstats, halloffame=hof, verbose=True)

    pool.close()

    # gen = log.select("gen") 
    # fit_maxs = log.chapters["fitness"].select("max")
    # best_fit = truncate(hof[0].fitness.values[0], 0)
    # nodes, edges, labels = gp.graph(hof[0])

    # append_to_excel=[]
    # append_to_excel.append(str(hof[0]))
    # append_to_excel.extend(fit_maxs)
    # print(append_to_excel)
    # write_to_excel(append_to_excel, str(GRAV), 'random_full_raw_data.xlsx')

    # Prints the fitness score of the best individual
    # print(best_fit)

    # Prints the individual's tree in string form
    # print(hof[0])

    # Graphs the fitness score of every ind over the generations and displays it
    # save_graph(gen, fit_maxs, best_fit)

    path_to_read='random_full_raw_data.xlsx'
    GRAV='-9.81'
    inds = get_one_column(path_to_read, GRAV, 'A')
    # Creates an env and displays the best ind being tested in the env
    fullObsEvalIndividual(inds[0], pset, -9.81, True)
    # fullObsEvalIndividual(hof[0], pset, -9.81, True)

    # return pop, log, hof

if __name__ == "__main__":
    main()
