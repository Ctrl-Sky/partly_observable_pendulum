import matplotlib.pyplot as plt
# import pygraphviz as pgv
# import pandas as pd

from openpyxl import load_workbook

# Takes an individual and plots it as a tree graph and saves it to a pdf
def plot_as_tree(nodes, edges, labels, best_fit):
    g = pgv.AGraph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    g.layout(prog="dot")

    for i in nodes:
        n = g.get_node(i)
        n.attr["label"] = labels[i]
    g.draw(str(best_fit)+".pdf")

# Get info on an individual and append it to a list
def best_ind_info(fit_mins, best_fit, hof, labels, ask):
    unused, used = find_unused_functions(labels)

    if ask == True:
        inp = input("Pass or fail?: ")
        notes = input("notes: ")
    else:
        inp = 'passed'
        notes = 'replace'

    fit_mins.append(best_fit)
    fit_mins.append(inp)
    if inp == 'passed':
        fit_mins.append(str(hof[0]))
    else:
        fit_mins.append(' ')
    fit_mins.append(unused)
    fit_mins.append(used)
    fit_mins.append(notes)

    return fit_mins

# Append the fitness information to an excel sheet
def write_to_excel(fit_mins, sheet_name, path):
    workbook = load_workbook(filename=path)

    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)


    workbook.active=workbook[sheet_name]

    workbook.active.append(fit_mins)

    workbook.save(filename=path)

def create_sheet(fit_mins, sheet_name, path):
    workbook = load_workbook(filename=path)

    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        workbook.active=workbook[sheet_name]
        workbook.active.append(fit_mins)
        workbook.save(filename=path)

# Creates and shows the graph of the fitness for then entire population
def plot_onto_graph(gen, fit_mins, best_fit):

    # Simply change the lines in quottation above to change the values you want to graph

    fig, ax1 = plt.subplots() # Allows you to create multiple plots in one figure
    line1 = ax1.plot(gen, fit_mins, 'b-', label="Maximum Fitness") # Plots using gen as x value and fit_mins as y, both are list
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Fitness", color="b")
    for tl in ax1.get_yticklabels(): # Changes colour of ticks and numbers on axis
        tl.set_color("b")

    lns = line1 # lns is a list containing both lines [line1, line2]
    labs = [l.get_label() for l in lns] # labs contains the labels of each line (Minimum Fitness and Average Size)
    ax1.legend(lns, labs, loc="lower right") # Adds then a legend

    plt.axis([min(gen), max(gen), min(fit_mins), 0])
    plt.show()

def save_graph(seed, gen, fit_mins, best_fit):
    colours = ["r-", "g-", "b-", "c-", "m-", "k-"]

    # Simply change the lines in quottation above to change the values you want to graph
    # Allows you to create multiple plots in one figure

    (fig, ax1) = plt.subplots()
    # Plots using gen as x value and fit_mins as y, both are list
    line1 = ax1.plot(
        gen, fit_mins, 'b-', label="Maximum Fitness"
    )
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Fitness", color="b")
    for (
        tl
    ) in ax1.get_yticklabels():  # Changes colour of ticks and numbers on axis
        tl.set_color("b")

    lns = line1  # lns is a list containing both lines [line1, line2]
    # labs contains the labels of each line (Minimum Fitness and Average Size)
    labs = [l.get_label() for l in lns]
    ax1.legend(lns, labs, loc="lower right")  # Adds then a legend

    plt.axis([min(gen), max(gen), min(fit_mins), 0])
    # plt.show()
    plt.savefig(str(seed) + "_fit_curve.png")

# Find used and unused functions in individual
def find_unused_functions(labels):
    used_functions = set(list(labels.values()))
    all_functions = {'add', 'conditional', 'ang_vel', 'sub', 'asin', 'acos', 'sin', 'cos', 'max', 'protectedDiv', 'limit', 'tan', 'atan', 'y1', 'y2', 'y3', 'x1', 'x2', 'x3'}
    unused_functions = all_functions.difference(used_functions)

    string1 = ''
    for i in unused_functions:
        string1 = string1 + i +', '

    string2 = ''
    for i in used_functions:
        string2 = string2 + i + ', '

    return string1, string2

def get_one_column(path, sheet, column_name):
    df = pd.read_excel(path, sheet, usecols=column_name)
    column_list = []
    for i in df.values:
        column_list.append(i[0])
    
    return column_list

