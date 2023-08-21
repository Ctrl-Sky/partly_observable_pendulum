# genetic programming for Gymnasium Pendulum task
# https://www.gymlibrary.dev/environments/classic_control/pendulum/ 

import numpy
import random
import gymnasium as gym
import operator
import matplotlib.pyplot as plt
import math
from openpyxl import load_workbook
import sys
import inspect
from inspect import isclass
from operator import attrgetter

from deap import algorithms
from deap import base
from deap import creator
from deap import tools
from deap import gp

import pygraphviz as pgv

# parallel
import multiprocessing

#parallel
import multiprocessing

# Import modules from different directory
import os
import sys
PATH=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(PATH)

from modules.prim_functions import *
from modules.output_functions import *
from modules.eval_individual import partObsEvalIndividual

GRAV=9.81
POP=2
GENS=2
TOURN_SIZE=5

# def write_to_excel(fit_mins, sheet_name, path):
#     workbook = load_workbook(filename=path)

#     if sheet_name not in workbook.sheetnames:
#         workbook.create_sheet(sheet_name)
#         workbook.active=workbook[sheet_name]
#         workbook.active.append(['ind', 'fitness'])


#     workbook.active=workbook[sheet_name]

#     workbook.active.append(fit_mins)

#     workbook.save(filename=path)

# def truncate(number, decimals=0):
#     if math.isinf(number) or math.isnan(number):
#         return 0
#     if not isinstance(decimals, int):
#         raise TypeError("decimal places must be an integer.")
#     elif decimals < 0:
#         raise ValueError("decimal places has to be 0 or more.")
#     elif decimals == 0:
#         return math.trunc(number)

#     factor = 10.0**decimals
#     num = number * factor
#     if math.isinf(num) or math.isnan(num):
#         return 0
#     return math.trunc(num) / factor

# def conditional(input1, input2):
#     if input1 < input2:
#         return -input1
#     else: return input1

# def limit(input, minimum, maximum):
#     if input < minimum:
#         return minimum
#     elif input > maximum:
#         return maximum
#     else:
#         return input

# def protectedDiv(left, right):
#     try:
#         return truncate(left, 8) / truncate(right, 8)
#     except ZeroDivisionError:
#         return 1

# def ang_vel(y2, y1, x2, x1):
#     top = atan(y2, x2) - atan(y1, x1)
#     bottom = y2 - y1
#     return protectedDiv(top, bottom)

# def acos(x, y):
#     if protectedDiv(x, y) < 1 and protectedDiv(x, y) > -1:
#         return math.acos(x/y)
#     elif protectedDiv(y, x) < 1 and protectedDiv(y, x) > -1:
#         return math.acos(y/x)
#     else:
#         return x

# def asin(x, y):
#     if protectedDiv(y, x) < 1 and protectedDiv(y, x) > -1:
#         return math.asin(y/x)
#     elif protectedDiv(y, x) < 1 and protectedDiv(y, x) > -1:
#         return math.asin(y/x)
#     else:
#         return x

# def atan(x, y):
#     return math.atan(protectedDiv(x, y))

def partObsEvalIndividual(individual, pset, grav, test=False):
    env_train = gym.make('Pendulum-v1', g=grav) # For training
    env_test = gym.make('Pendulum-v1', g=grav, render_mode="human") # For rendering the best one
    env = env_train
    num_episode = 30 # Basically the amount of simulations ran
    if test:
        env = env_test
        num_episode = 1
    
    # Transform the tree expression to functional Python code
    get_action = gp.compile(individual, pset)
    fitness = 0
    failed = False
    for x in range(0, num_episode):
        done = False
        truncated = False
        observation = env.reset() # Reset the pole to some random location and defines the things in observation
        observation = observation[0]
        episode_reward = 0
        num_steps = 0
        max_steps = 300
        timeout = False

        prev_y = observation[0]
        prev_x = observation[1]
        last_y = observation[0]
        last_x = observation[1]

        while not (done or timeout):
            if failed:
                action = 0
            else:
                # use the tree to compute action, plugs values of observation into get_action
                                    
                if num_steps == 0:
                    action = get_action(observation[0], observation[1], prev_y, prev_x, last_y, last_x)
                    prev_y = observation[0]
                    prev_x = observation[1]
                else:
                    action = get_action(observation[0], observation[1], prev_y, prev_x, last_y, last_x)
                    temp_y = prev_y
                    temp_x = prev_x
                    prev_y = observation[0]
                    prev_x = observation[1]
                    last_y = temp_y
                    last_x = temp_x
                # action = get_action(observation[0], observation[1], observation[2])
                
                action = (action, )

            try: observation, reward, done, truncated, info = env.step(action) # env.step will return the new observation, reward, done, truncated, info
            except:
                failed = True
                observation, reward, done, truncated, info = env.step(0)
            episode_reward += reward

            num_steps += 1
            if num_steps >= max_steps:
                timeout = True
            
        fitness += episode_reward

    fitness = fitness/num_episode        
    return (0,) if failed else (fitness,)

# Set up primitives and terminals
pset = gp.PrimitiveSet("MAIN", 6)
pset.addPrimitive(operator.add, 2)
pset.addPrimitive(conditional, 2)
pset.addPrimitive(ang_vel, 4)
pset.addPrimitive(protectedDiv, 2)
pset.addPrimitive(operator.sub, 2)
pset.addPrimitive(asin, 2)
pset.addPrimitive(acos, 2)
pset.addPrimitive(atan, 2)
pset.addPrimitive(math.cos, 1)
pset.addPrimitive(math.sin, 1)
pset.addPrimitive(math.tan, 1)
pset.addPrimitive(max, 2)
pset.addPrimitive(limit, 3)

pset.renameArguments(ARG0='y1')
pset.renameArguments(ARG1='x1')
pset.renameArguments(ARG2='y2')
pset.renameArguments(ARG3='x2')
pset.renameArguments(ARG4='y3')
pset.renameArguments(ARG5='x3')

# Prepare individual and pendulum
creator.create("FitnessMax", base.Fitness, weights=(1.0,))
creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)

toolbox = base.Toolbox()
toolbox.register("expr", gp.genFull, pset=pset, min_=2, max_=3)
toolbox.register("individual", tools.initIterate, creator.Individual,
                 toolbox.expr)
toolbox.register("population", tools.initRepeat, list, toolbox.individual)

# Register functions in the toolbox needed for evolution
toolbox.register("evaluate", partObsEvalIndividual, pset=pset, grav=GRAV)
toolbox.register("select", tools.selTournament, tournsize=TOURN_SIZE)
toolbox.register("mate", gp.cxOnePoint)
toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=17))
toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=17))

def main():
    pop = toolbox.population(n=POP)
    hof = tools.HallOfFame(1)

    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", numpy.mean)
    mstats.register("std", numpy.std)
    mstats.register("min", numpy.min)
    mstats.register("max", numpy.max)

    pool = multiprocessing.Pool(processes=2) # parllel (Process Pool of 16 workers)
    toolbox.register("map", pool.map) # parallel

    pop, log = algorithms.eaSimple(pop, toolbox, 0.2, 0.5, GENS, stats=mstats, halloffame=hof, verbose=True)

    pool.close() # parallel
    
    gen = log.select("gen") 
    fit_mins = log.chapters["fitness"].select("max")
    best_fit = truncate(hof[0].fitness.values[0], 0)
    nodes, edges, labels = gp.graph(hof[0])

    create_sheet(['ind', 'fitness'], str(GRAV), 'part_obs_raw_data.xlsx')
    append_to_excel=[]
    append_to_excel.append(str(hof[0]))
    append_to_excel.append(best_fit)
    write_to_excel(append_to_excel, str(GRAV), 'part_obs_raw_data.xlsx')

    # Prints the fitness score of the best individual
    # print(best_fit)

    # Prints the individual's tree in string form
    # print(hof[0])

    # Graphs the fitness score of every ind over the generations and displays it
    # plot_onto_graph(gen, fit_mins, best_fit)

    # Creates an env and displays the best ind being tested in the env
    # partObsEvalIndividual(hof[0], pset, 9.81, True)

    # Ask for relevant info about the best individual and then writes it
    # to a part_obs_training_data
    # fit_mins = best_ind_info(fit_mins, best_fit, hof, labels, ask=False)
    # write_to_excel(fit_mins, str(GRAV), "part_obs_raw_data.xlsx")

    return pop, log, hof

if __name__ == "__main__":
    main()
