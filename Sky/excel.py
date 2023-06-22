from openpyxl import load_workbook

workbook = load_workbook(filename="/Users/sky/Documents/Book1.xlsx")
sheet = workbook.active

data = ["Item 5", "Item 6", "Item 7", "Item 8"]

starting_row = 2 # Append after the last row of existing data
# starting_row = 10  # Specify a specific row to append the new data

column = sheet.max_column + 1

for row, value in enumerate(data, start=starting_row):
    sheet.cell(row=row, column=column, value=value)

workbook.save(filename="/Users/sky/Documents/Book1.xlsx")